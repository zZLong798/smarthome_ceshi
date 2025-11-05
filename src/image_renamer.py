#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
图片重命名器模块
根据PDID和设备简称重命名图片文件
"""

import os
import re
import shutil
from typing import Dict, List, Optional, Tuple
import openpyxl
from pypinyin import pinyin, Style


class ImageRenamer:
    """图片重命名器类"""
    
    def __init__(self, excel_file_path: str, image_dir: str):
        """
        初始化重命名器
        
        Args:
            excel_file_path: Excel文件路径
            image_dir: 图片目录路径
        """
        self.excel_file_path = excel_file_path
        self.image_dir = image_dir
        self.pdid_mapping: Dict[int, str] = {}  # 行号到PDID的映射
        self.device_name_mapping: Dict[int, str] = {}  # 行号到设备简称的映射
    
    def load_excel_data(self) -> bool:
        """
        加载Excel数据，获取PDID和设备简称映射
        
        Returns:
            bool: 是否成功加载
        """
        try:
            wb = openpyxl.load_workbook(self.excel_file_path)
            ws = wb.active
            
            # 找到关键列索引
            pdid_col = None
            device_name_col = None
            device_short_col = None
            
            # 扫描表头
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    header_str = str(header)
                    if '产品ID' in header_str or 'PDID' in header_str:
                        pdid_col = col
                    elif '设备名称' in header_str:
                        device_name_col = col
                    elif '设备简称' in header_str:
                        device_short_col = col
            
            print(f"关键列索引: PDID={pdid_col}, 设备名称={device_name_col}, 设备简称={device_short_col}")
            
            if not pdid_col:
                print("未找到产品ID列")
                return False
            
            # 加载数据映射
            for row in range(2, ws.max_row + 1):
                # 获取PDID
                pdid_cell = ws.cell(row=row, column=pdid_col)
                if pdid_cell.value:
                    pdid = str(pdid_cell.value).strip()
                    self.pdid_mapping[row] = pdid
                
                # 获取设备简称（优先使用设备简称列，如果没有则使用设备名称）
                device_short_name = None
                if device_short_col:
                    short_cell = ws.cell(row=row, column=device_short_col)
                    if short_cell.value:
                        device_short_name = str(short_cell.value).strip()
                
                if not device_short_name and device_name_col:
                    name_cell = ws.cell(row=row, column=device_name_col)
                    if name_cell.value:
                        device_short_name = str(name_cell.value).strip()
                
                if device_short_name:
                    self.device_name_mapping[row] = device_short_name
            
            print(f"加载Excel数据成功: PDID映射={len(self.pdid_mapping)}, 设备名称映射={len(self.device_name_mapping)}")
            return True
            
        except Exception as e:
            print(f"加载Excel数据失败: {e}")
            return False
    
    def chinese_to_pinyin(self, chinese_text: str) -> str:
        """
        将中文转换为拼音
        
        Args:
            chinese_text: 中文文本
            
        Returns:
            str: 拼音字符串
        """
        try:
            # 如果文本为空，返回默认值
            if not chinese_text:
                return "device"
            
            # 处理混合文本（中文+数字+字母）
            pinyin_result = []
            
            # 逐个字符处理
            for char in chinese_text:
                if '\u4e00' <= char <= '\u9fff':  # 中文字符
                    # 转换为拼音
                    char_pinyin = pinyin(char, style=Style.NORMAL)
                    if char_pinyin:
                        pinyin_result.append(char_pinyin[0][0])
                elif char.isalnum():  # 字母或数字
                    pinyin_result.append(char.lower())
                elif char in (' ', '-', '_'):  # 分隔符
                    pinyin_result.append('_')
                else:
                    # 其他字符（如中文标点符号）跳过
                    continue
            
            # 合并拼音结果
            pinyin_str = ''.join(pinyin_result)
            
            # 如果转换失败或结果为空，使用备选方案
            if not pinyin_str:
                simplified = re.sub(r'[^a-zA-Z0-9]', '_', chinese_text)
                pinyin_str = simplified.lower()
            
            # 清理结果：移除多余的下划线，确保格式规范
            pinyin_str = re.sub(r'_+', '_', pinyin_str)
            pinyin_str = pinyin_str.strip('_')
            
            # 如果结果仍然包含中文字符，使用更严格的转换
            if re.search(r'[\u4e00-\u9fff]', pinyin_str):
                print(f"警告：拼音转换结果仍包含中文字符: {chinese_text} -> {pinyin_str}")
                # 使用备选方案：移除所有非ASCII字符
                pinyin_str = re.sub(r'[^a-zA-Z0-9_]', '_', chinese_text)
                pinyin_str = re.sub(r'_+', '_', pinyin_str)
                pinyin_str = pinyin_str.strip('_').lower()
            
            return pinyin_str
            
        except Exception as e:
            print(f"中文转拼音失败 {chinese_text}: {e}")
            # 返回简化版本作为备选
            simplified = re.sub(r'[^a-zA-Z0-9]', '_', chinese_text)
            return simplified.lower()
    
    def sanitize_filename(self, filename: str) -> str:
        """
        清理文件名，移除非法字符
        
        Args:
            filename: 原始文件名
            
        Returns:
            str: 清理后的文件名
        """
        # 移除或替换非法字符
        invalid_chars = r'[<>:"/\\|?*]'
        sanitized = re.sub(invalid_chars, '_', filename)
        
        # 移除多余的下划线
        sanitized = re.sub(r'_+', '_', sanitized)
        
        # 移除首尾的下划线
        sanitized = sanitized.strip('_')
        
        return sanitized
    
    def generate_image_filename(self, pdid: str, device_name: str) -> str:
        """
        生成图片文件名
        
        Args:
            pdid: 产品ID
            device_name: 设备名称
            
        Returns:
            str: 生成的图片文件名
        """
        try:
            # 清理PDID
            clean_pdid = self.sanitize_filename(pdid)
            
            # 将设备名称转换为拼音
            pinyin_name = self.chinese_to_pinyin(device_name)
            
            # 生成文件名: pdid{id}_{设备简称拼音}.png
            filename = f"pdid{clean_pdid}_{pinyin_name}.png"
            
            print(f"生成文件名: PDID={clean_pdid}, 设备={device_name} -> {filename}")
            return filename
            
        except Exception as e:
            print(f"生成图片文件名失败: PDID={pdid}, 设备={device_name}, 错误={e}")
            # 备选方案
            return f"pdid{pdid}_device.png"
    
    def map_images_to_rows(self, image_files: List[str]) -> Dict[int, str]:
        """
        将图片文件映射到Excel行号（使用增强版映射功能）
        
        Args:
            image_files: 图片文件列表
            
        Returns:
            Dict[int, str]: 行号到图片文件路径的映射
        """
        row_image_mapping = {}
        
        try:
            # 使用增强版映射器建立完整的映射关系
            from src.enhanced_excel_image_mapper import EnhancedExcelImageMapper
            mapper = EnhancedExcelImageMapper(self.excel_file_path)
            
            # 解析增强版映射关系
            enhanced_mapping = mapper.parse_enhanced_mapping()
            
            if not enhanced_mapping:
                print("警告：增强版映射解析失败，回退到基础映射")
                return self._map_images_to_rows_fallback(image_files)
            
            # 构建图片ID到行号的映射
            image_id_to_row = {}
            for mapping_key, mapping_info in enhanced_mapping.items():
                if 'image_id' in mapping_info and 'row_number' in mapping_info:
                    image_id_to_row[mapping_info['image_id']] = mapping_info['row_number']
            
            # 通过图片ID映射
            mapped_count = 0
            for image_file in image_files:
                # 提取图片ID（文件名去掉扩展名）
                image_id = os.path.splitext(image_file)[0]
                
                # 查找对应的行号
                if image_id in image_id_to_row:
                    row_num = image_id_to_row[image_id]
                    row_image_mapping[row_num] = os.path.join(self.image_dir, image_file)
                    mapped_count += 1
                    print(f"图片映射(增强版): 行{row_num} -> {image_file}")
                else:
                    print(f"无法为图片 {image_file} 找到对应的Excel行")
            
            print(f"图片映射完成: 共映射 {mapped_count} 张图片 (增强版映射)")
            return row_image_mapping
            
        except Exception as e:
            print(f"映射图片到行号失败: {e}")
            # 降级到基础映射
            return self._map_images_to_rows_fallback(image_files)
    
    def _map_images_to_rows_fallback(self, image_files: List[str]) -> Dict[int, str]:
        """
        降级映射方法（使用基础映射）
        
        Args:
            image_files: 图片文件列表
            
        Returns:
            Dict[int, str]: 行号到图片文件路径的映射
        """
        row_image_mapping = {}
        
        try:
            # 解析DISPIMG公式，获取行号到图片ID的映射
            from src.excel_image_extractor import ExcelImageExtractor
            extractor = ExcelImageExtractor(self.excel_file_path)
            dispimg_mapping = extractor.parse_dispimg_formulas(self.excel_file_path)
            
            # 反转映射：图片ID到行号的映射
            image_id_to_row = {}
            for row, image_id in dispimg_mapping.items():
                image_id_to_row[image_id] = row
            
            # 通过图片ID映射
            mapped_count = 0
            for image_file in image_files:
                # 提取图片ID（文件名去掉扩展名）
                image_id = os.path.splitext(image_file)[0]
                
                # 查找对应的行号
                if image_id in image_id_to_row:
                    row_num = image_id_to_row[image_id]
                    row_image_mapping[row_num] = os.path.join(self.image_dir, image_file)
                    mapped_count += 1
                    print(f"图片映射(降级): 行{row_num} -> {image_file}")
                else:
                    print(f"无法为图片 {image_file} 找到对应的Excel行")
            
            print(f"图片映射完成(降级): 共映射 {mapped_count} 张图片")
            return row_image_mapping
            
        except Exception as e:
            print(f"降级映射失败: {e}")
            return {}
    
    def rename_images(self, image_mapping: Dict[str, str]) -> Dict[str, str]:
        """
        重命名图片文件（使用增强版映射功能）
        
        Args:
            image_mapping: 图片ID到文件路径的映射
            
        Returns:
            Dict[str, str]: 新文件名到原文件路径的映射
        """
        renamed_images = {}
        
        try:
            # 首先加载Excel数据
            if not self.load_excel_data():
                print("无法加载Excel数据，无法重命名图片")
                return {}
            
            # 创建重命名后的目录（保存到images目录）
            # 注意：这里只生成新文件名映射，不实际保存文件
            # 实际保存由ImageSaver负责
            
            # 使用增强版映射器建立完整的映射关系
            from src.enhanced_excel_image_mapper import EnhancedExcelImageMapper
            mapper = EnhancedExcelImageMapper(self.excel_file_path)
            
            # 解析增强版映射关系
            enhanced_mapping = mapper.parse_enhanced_mapping()
            
            if not enhanced_mapping:
                print("警告：增强版映射解析失败，回退到基础映射")
                return self._rename_images_fallback(image_mapping, renamed_dir)
            
            # 构建图片ID到行号的映射
            image_id_to_row = {}
            for mapping_key, mapping_info in enhanced_mapping.items():
                if 'image_id' in mapping_info and 'row_number' in mapping_info:
                    image_id_to_row[mapping_info['image_id']] = mapping_info['row_number']
            
            # 重命名图片
            renamed_count = 0
            for image_id, image_path in image_mapping.items():
                # 查找对应的行号
                matching_row = None
                
                # 方法1: 通过增强版映射查找
                if image_id in image_id_to_row:
                    matching_row = image_id_to_row[image_id]
                
                # 方法2: 尝试从图片ID中提取行号（备选方案）
                if not matching_row:
                    # 尝试从图片ID中提取数字作为行号
                    import re
                    match = re.search(r'\d+', image_id)
                    if match:
                        row_num = int(match.group())
                        if row_num >= 2 and row_num <= len(self.pdid_mapping) + 2:
                            matching_row = row_num
                
                if matching_row:
                    pdid = self.pdid_mapping.get(matching_row)
                    device_name = self.device_name_mapping.get(matching_row)
                    
                    if pdid and device_name:
                        # 生成新文件名
                        new_filename = self.generate_image_filename(pdid, device_name)
                        
                        # 只生成映射关系，不实际复制文件
                        renamed_images[new_filename] = image_path
                        renamed_count += 1
                        
                        print(f"重命名图片: 行{matching_row} -> {os.path.basename(image_path)} -> {new_filename}")
                    else:
                        print(f"无法为图片 {image_id} 找到对应的PDID或设备名称 (行{matching_row})")
                else:
                    print(f"无法为图片 {image_id} 找到对应的Excel行")
            
            print(f"图片重命名完成: 共重命名 {renamed_count} 张图片 (增强版映射)")
            return renamed_images
            
        except Exception as e:
            print(f"重命名图片失败: {e}")
            # 降级到基础映射
            return self._rename_images_fallback(image_mapping, renamed_dir)
    
    def _rename_images_fallback(self, image_mapping: Dict[str, str], renamed_dir: str) -> Dict[str, str]:
        """
        降级重命名方法（使用基础映射）
        
        Args:
            image_mapping: 图片ID到文件路径的映射
            renamed_dir: 重命名后的目录
            
        Returns:
            Dict[str, str]: 新文件名到原文件路径的映射
        """
        renamed_images = {}
        
        try:
            # 解析DISPIMG公式，获取行号到图片ID的映射
            from src.excel_image_extractor import ExcelImageExtractor
            extractor = ExcelImageExtractor(self.excel_file_path)
            dispimg_mapping = extractor.parse_dispimg_formulas(self.excel_file_path)
            
            # 反转映射：图片ID到行号的映射
            image_id_to_row = {}
            for row, image_id in dispimg_mapping.items():
                image_id_to_row[image_id] = row
            
            # 重命名图片
            renamed_count = 0
            for image_id, image_path in image_mapping.items():
                # 查找对应的行号
                matching_row = None
                
                # 通过DISPIMG公式映射查找
                if image_id in image_id_to_row:
                    matching_row = image_id_to_row[image_id]
                
                if matching_row:
                    pdid = self.pdid_mapping.get(matching_row)
                    device_name = self.device_name_mapping.get(matching_row)
                    
                    if pdid and device_name:
                        # 生成新文件名
                        new_filename = self.generate_image_filename(pdid, device_name)
                        
                        # 只生成映射关系，不实际复制文件
                        renamed_images[new_filename] = image_path
                        renamed_count += 1
                        
                        print(f"重命名图片(降级): 行{matching_row} -> {os.path.basename(image_path)} -> {new_filename}")
                    else:
                        print(f"无法为图片 {image_id} 找到对应的PDID或设备名称 (行{matching_row})")
                else:
                    print(f"无法为图片 {image_id} 找到对应的Excel行")
            
            print(f"图片重命名完成(降级): 共重命名 {renamed_count} 张图片")
            return renamed_images
            
        except Exception as e:
            print(f"降级重命名失败: {e}")
            return {}
    
    def get_renaming_mapping(self) -> Dict[int, Tuple[str, str]]:
        """
        获取重命名映射关系
        
        Returns:
            Dict[int, Tuple[str, str]]: 行号到(PDID, 设备名称)的映射
        """
        mapping = {}
        
        try:
            if not self.load_excel_data():
                return {}
            
            for row in range(2, len(self.pdid_mapping) + 2):
                if row in self.pdid_mapping and row in self.device_name_mapping:
                    pdid = self.pdid_mapping[row]
                    device_name = self.device_name_mapping[row]
                    mapping[row] = (pdid, device_name)
            
            return mapping
            
        except Exception as e:
            print(f"获取重命名映射失败: {e}")
            return {}


def main():
    """测试函数"""
    # 测试图片重命名
    excel_file = "智能家居模具库.xlsx"
    image_dir = "extracted_images"
    
    if os.path.exists(excel_file) and os.path.exists(image_dir):
        renamer = ImageRenamer(excel_file, image_dir)
        
        # 测试重命名映射
        mapping = renamer.get_renaming_mapping()
        print(f"重命名映射: {len(mapping)} 个映射关系")
        
        # 测试文件名生成
        test_cases = [
            ("PD001", "智能开关"),
            ("PD002", "智能灯带"),
            ("PD003", "吸顶AP")
        ]
        
        for pdid, device_name in test_cases:
            filename = renamer.generate_image_filename(pdid, device_name)
            print(f"测试文件名生成: {pdid} + {device_name} -> {filename}")
    else:
        print("测试文件不存在")


if __name__ == "__main__":
    main()