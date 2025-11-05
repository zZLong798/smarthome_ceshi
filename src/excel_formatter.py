"""
Excel格式美化模块 - 负责采购清单的格式美化
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelFormatter:
    """Excel格式美化器类"""
    
    def __init__(self):
        """初始化格式美化器"""
        # 定义颜色方案 - 橙色/黄色主题
        self.colors = {
            'header_bg': 'FFD2691E',  # 深橙色背景
            'header_font': 'FFFFFFFF',  # 白色字体
            'summary_bg': 'FFFFD700',   # 金黄色背景
            'summary_font': 'FF000000', # 黑色字体
            'even_row_bg': 'FFFFF8DC',  # 浅黄色背景（偶数行）
            'odd_row_bg': 'FFFFFACD',   # 浅橙色背景（奇数行）
            'data_font': 'FF000000',    # 黑色字体
            'border': 'FF000000'        # 黑色边框
        }
        
        # 定义字体
        self.fonts = {
            'header': Font(name='微软雅黑', size=12, bold=True, color=self.colors['header_font']),
            'summary': Font(name='微软雅黑', size=11, bold=True, color=self.colors['summary_font']),
            'data': Font(name='微软雅黑', size=10, color=self.colors['data_font']),
            'link': Font(name='微软雅黑', size=10, color='FF0000FF', underline='single')  # 蓝色链接
        }
        
        # 定义填充样式
        self.fills = {
            'header': PatternFill(start_color=self.colors['header_bg'], end_color=self.colors['header_bg'], fill_type='solid'),
            'summary': PatternFill(start_color=self.colors['summary_bg'], end_color=self.colors['summary_bg'], fill_type='solid'),
            'even_row': PatternFill(start_color=self.colors['even_row_bg'], end_color=self.colors['even_row_bg'], fill_type='solid'),
            'odd_row': PatternFill(start_color=self.colors['odd_row_bg'], end_color=self.colors['odd_row_bg'], fill_type='solid')
        }
        
        # 定义边框
        thin_border = Side(border_style='thin', color=self.colors['border'])
        self.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
        
        # 定义对齐方式
        self.alignments = {
            'center': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'left': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'right': Alignment(horizontal='right', vertical='center', wrap_text=True)
        }
    
    def format_header_row(self, worksheet, header_row=1):
        """
        格式化标题行
        
        Args:
            worksheet: Excel工作表对象
            header_row: 标题行号（默认为1）
        """
        max_col = worksheet.max_column
        
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=header_row, column=col)
            
            # 设置字体和填充
            cell.font = self.fonts['header']
            cell.fill = self.fills['header']
            cell.alignment = self.alignments['center']
            cell.border = self.border
            
            # 自动调整列宽
            column_letter = get_column_letter(col)
            worksheet.column_dimensions[column_letter].width = 15
    
    def format_data_rows(self, worksheet, start_row=2):
        """
        格式化数据行
        
        Args:
            worksheet: Excel工作表对象
            start_row: 数据开始行号（默认为2）
        """
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        for row in range(start_row, max_row + 1):
            # 判断是否为汇总行（通常包含"汇总"或"总计"字样）
            is_summary_row = False
            for col in range(1, max_col + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value and ('汇总' in str(cell_value) or '总计' in str(cell_value)):
                    is_summary_row = True
                    break
            
            if is_summary_row:
                # 格式化汇总行
                self.format_summary_row(worksheet, row)
            else:
                # 格式化普通数据行
                row_parity = 'even' if row % 2 == 0 else 'odd'
                
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    
                    # 设置字体和填充
                    cell.font = self.fonts['data']
                    cell.fill = self.fills[f'{row_parity}_row']
                    cell.border = self.border
                    
                    # 根据列类型设置对齐方式
                    if col in [1, 2, 3]:  # 产品图片、产品名称、品牌
                        cell.alignment = self.alignments['center']
                    elif col in [4, 5, 6]:  # 数量、单价、小计
                        cell.alignment = self.alignments['right']
                    else:  # 其他列
                        cell.alignment = self.alignments['left']
                    
                    # 设置自动换行
                    cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical,
                        wrap_text=True
                    )
    
    def format_summary_row(self, worksheet, row):
        """
        格式化汇总行
        
        Args:
            worksheet: Excel工作表对象
            row: 汇总行号
        """
        max_col = worksheet.max_column
        
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            
            # 设置字体和填充
            cell.font = self.fonts['summary']
            cell.fill = self.fills['summary']
            cell.border = self.border
            
            # 设置对齐方式
            if col in [4, 5, 6]:  # 数量、单价、小计列
                cell.alignment = self.alignments['right']
            else:
                cell.alignment = self.alignments['center']
    
    def adjust_column_widths(self, worksheet):
        """
        自动调整列宽以适应内容
        
        Args:
            worksheet: Excel工作表对象
        """
        # 设置固定列宽
        column_widths = {
            'A': 2.0,   # 产品图片列（0.9cm图片）
            'B': 20.0,  # 产品名称
            'C': 12.0,  # 品牌
            'D': 8.0,   # 数量
            'E': 10.0,  # 单价
            'F': 12.0,  # 小计
            'G': 25.0,  # 产品链接
            'H': 15.0   # 备注
        }
        
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        # 设置行高
        worksheet.row_dimensions[1].height = 25  # 标题行高度
        for row in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 60  # 数据行高度（适应图片和换行文本）
    
    def format_worksheet(self, worksheet):
        """
        完整格式化工作表
        
        Args:
            worksheet: Excel工作表对象
        """
        # 调整列宽和行高
        self.adjust_column_widths(worksheet)
        
        # 格式化标题行
        self.format_header_row(worksheet)
        
        # 格式化数据行
        self.format_data_rows(worksheet)
        
        # 设置打印区域
        worksheet.print_area = f'A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}'
        
        # 设置页面布局
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
    
    def format_hyperlink_cells(self, worksheet):
        """
        专门格式化超链接单元格
        
        Args:
            worksheet: Excel工作表对象
        """
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                
                # 检查单元格是否包含超链接
                if cell.hyperlink:
                    cell.font = self.fonts['link']


def test_excel_formatter():
    """测试Excel格式美化器"""
    from openpyxl import Workbook
    
    # 创建测试工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "测试格式"
    
    # 添加测试数据
    headers = ["产品图片", "产品名称", "品牌", "数量", "单价", "小计", "产品链接", "备注"]
    data = [
        ["", "领普二键智能开关", "领普", 3, 89, 267, "https://example.com", "智能家居必备"],
        ["", "易来四键智能开关", "易来", 4, 115, 460, "https://example.com", "高品质开关"],
        ["", "汇总", "", 7, "", 727, "", "总计"]
    ]
    
    # 添加标题行
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 添加数据行
    for row, row_data in enumerate(data, 2):
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # 应用格式美化
    formatter = ExcelFormatter()
    formatter.format_worksheet(ws)
    
    # 保存测试文件
    wb.save("test_formatting.xlsx")
    print("格式美化测试完成，已保存为 test_formatting.xlsx")


if __name__ == "__main__":
    test_excel_formatter()