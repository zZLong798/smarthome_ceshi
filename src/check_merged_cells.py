#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
检查模板中的合并单元格
"""

import openpyxl

def check_merged_cells():
    """检查模板中的合并单元格"""
    
    print("🔍 检查模板中的合并单元格...")
    
    try:
        # 加载模板文件
        workbook = openpyxl.load_workbook('../采购清单模板.xlsx')
        worksheet = workbook.active
        
        print(f"✅ 模板文件加载成功")
        print(f"📊 合并单元格数量: {len(worksheet.merged_cells.ranges)}")
        
        # 显示所有合并单元格范围
        print("\n📐 合并单元格范围:")
        for merged_range in worksheet.merged_cells.ranges:
            print(f"   {merged_range}")
            
            # 检查合并单元格的内容
            top_left_cell = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
            print(f"     内容: '{top_left_cell.value}'")
            print(f"     位置: 行{merged_range.min_row}-{merged_range.max_row}, 列{merged_range.min_col}-{merged_range.max_col}")
        
        # 检查数据行范围
        print(f"\n📏 数据行范围: 2-{worksheet.max_row}")
        
        # 检查哪些行有合并单元格
        print("\n🔍 检查数据行中的合并单元格:")
        for row in range(2, min(10, worksheet.max_row + 1)):  # 只检查前几行
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                # 检查单元格是否在合并范围内
                for merged_range in worksheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        print(f"   行{row}列{col}在合并范围内: {merged_range}")
                        break
        
        workbook.close()
        print("\n🎉 合并单元格检查完成")
        
    except Exception as e:
        print(f"❌ 检查过程中出错: {e}")

def main():
    """主函数"""
    print("=" * 50)
    print("📋 模板合并单元格分析")
    print("=" * 50)
    
    check_merged_cells()

if __name__ == "__main__":
    main()