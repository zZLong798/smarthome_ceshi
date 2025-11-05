#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel图片提取器模块
专门用于从Excel文件中提取嵌入图片和解析DISPIMG公式
"""
import os
import shutil
import tempfile
import zipfile
import xml.etree.ElementTree as ET
from typing import Dict, List, Any, Optional, Tuple
import re
import openpyxl


class ExcelImageExtractor:
    """Excel图片提取器类"""
    
    def __init__(self, excel_file_path: str, temp_dir: str = "temp_excel_extract"):
        """
        初始化图片提取器
        
        Args:
            excel_file_path: Excel文件路径
            temp_dir: 临时解压目录
        """
        self.excel_file_path = excel_file_path
        self.temp_dir = temp_dir
        self.image_mapping: Dict[str, str] = {}  # 图片ID到文件名的映射
        
        # 确保临时目录存在
        if not os.path.exists(self.temp_dir):
            os.makedirs(self.temp_dir)
    
    def copy_excel_as_zip(self) -> str:
        """
        复制Excel文件为zip格式
        
        Returns:
            str: 复制的zip文件路径
        """
        zip_file_path = os.path.join(self.temp_dir, "excel_copy.zip")
        
        try:
            # 复制Excel文件
            shutil.copy2(self.excel_file_path, zip_file_path)
            print(f"Excel文件已复制为zip格式: {zip_file_path}")
            return zip_file_path
        except Exception as e:
            print(f"复制Excel文件失败: {e}")
            raise
    
    def extract_zip_structure(self, zip_file_path: str) -> str:
        """
        解压zip文件结构
        
        Args:
            zip_file_path: zip文件路径
            
        Returns:
            str: 解压后的目录路径
        """
        extract_dir = os.path.join(self.temp_dir, "extracted")
        
        try:
            # 确保解压目录存在
            if os.path.exists(extract_dir):
                shutil.rmtree(extract_dir)
            os.makedirs(extract_dir)
            
            # 解压zip文件
            with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
                zip_ref.extractall(extract_dir)
            
            print(f"Excel结构已解压到: {extract_dir}")
            return extract_dir
        except Exception as e:
            print(f"解压Excel结构失败: {e}")
            raise
    
    def parse_image_mappings(self, extract_dir: str) -> Dict[str, str]:
        """
        解析图片映射关系
        
        Args:
            extract_dir: 解压后的目录路径
            
        Returns:
            Dict[str, str]: 图片ID到文件名的映射
        """
        image_mapping = {}
        
        try:
            # 解析xl/media目录中的图片文件
            media_dir = os.path.join(extract_dir, "xl", "media")
            if os.path.exists(media_dir):
                for filename in os.listdir(media_dir):
                    if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
                        # 提取图片ID（文件名去掉扩展名）
                        image_id = os.path.splitext(filename)[0]
                        image_mapping[image_id] = filename
                        print(f"发现图片映射: ID={image_id}, 文件={filename}")
            
            # 解析xl/drawings目录中的drawing XML文件
            drawings_dir = os.path.join(extract_dir, "xl", "drawings")
            if os.path.exists(drawings_dir):
                for drawing_file in os.listdir(drawings_dir):
                    if drawing_file.endswith('.xml'):
                        drawing_path = os.path.join(drawings_dir, drawing_file)
                        self._parse_drawing_xml(drawing_path, image_mapping)
            
            # 解析worksheets目录中的sheet XML文件
            worksheets_dir = os.path.join(extract_dir, "xl", "worksheets")
            if os.path.exists(worksheets_dir):
                for sheet_file in os.listdir(worksheets_dir):
                    if sheet_file.endswith('.xml'):
                        sheet_path = os.path.join(worksheets_dir, sheet_file)
                        self._parse_sheet_xml(sheet_path, image_mapping)
            
            print(f"解析到 {len(image_mapping)} 个图片映射关系")
            return image_mapping
            
        except Exception as e:
            print(f"解析图片映射失败: {e}")
            return {}
    
    def _parse_drawing_xml(self, drawing_path: str, image_mapping: Dict[str, str]):
        """解析drawing XML文件"""
        try:
            tree = ET.parse(drawing_path)
            root = tree.getroot()
            
            # 命名空间处理
            namespaces = {
                'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'
            }
            
            # 查找所有图片引用
            for pic in root.findall('.//xdr:pic', namespaces):
                blip = pic.find('.//a:blip', namespaces)
                if blip is not None:
                    embed_id = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                    if embed_id:
                        # 这里可以记录图片引用关系
                        pass
                        
        except Exception as e:
            print(f"解析drawing XML失败 {drawing_path}: {e}")
    
    def _parse_sheet_xml(self, sheet_path: str, image_mapping: Dict[str, str]):
        """解析sheet XML文件"""
        try:
            tree = ET.parse(sheet_path)
            root = tree.getroot()
            
            # 命名空间处理
            namespaces = {
                '': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
            }
            
            # 查找图片引用
            for drawing in root.findall('.//drawing', namespaces):
                # 这里可以记录图片与单元格的关系
                pass
                
        except Exception as e:
            print(f"解析sheet XML失败 {sheet_path}: {e}")
    
    def extract_images_from_media(self, extract_dir: str, output_dir: str) -> Dict[str, str]:
        """
        从media目录提取图片到输出目录
        
        Args:
            extract_dir: 解压后的目录路径
            output_dir: 输出目录路径
            
        Returns:
            Dict[str, str]: 图片ID到输出文件路径的映射
        """
        extracted_images = {}
        
        try:
            # 确保输出目录存在
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            # 提取xl/media目录中的图片
            media_dir = os.path.join(extract_dir, "xl", "media")
            if os.path.exists(media_dir):
                for filename in os.listdir(media_dir):
                    if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
                        source_path = os.path.join(media_dir, filename)
                        target_path = os.path.join(output_dir, filename)
                        
                        # 复制图片文件
                        shutil.copy2(source_path, target_path)
                        
                        # 记录映射关系
                        image_id = os.path.splitext(filename)[0]
                        extracted_images[image_id] = target_path
                        print(f"图片已提取: {filename} -> {target_path}")
            
            print(f"成功提取 {len(extracted_images)} 张图片")
            return extracted_images
            
        except Exception as e:
            print(f"提取图片失败: {e}")
            return {}
    
    def parse_dispimg_formulas(self, excel_file_path: str) -> Dict[int, str]:
        """
        解析Excel中的DISPIMG公式，提取图片ID
        
        Args:
            excel_file_path: Excel文件路径
            
        Returns:
            Dict[int, str]: 行号到图片ID的映射
        """
        dispimg_mapping = {}
        
        try:
            # 使用openpyxl读取Excel
            wb = openpyxl.load_workbook(excel_file_path, data_only=False)
            ws = wb.active
            
            # 找到设备图片列
            image_col_index = None
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header and '图片' in str(header):
                    image_col_index = col
                    print(f"找到设备图片列: 第{col}列 - {header}")
                    break
            
            if not image_col_index:
                print("未找到设备图片列")
                return {}
            
            # 解析所有行的DISPIMG公式
            for row in range(2, ws.max_row + 1):
                image_cell = ws.cell(row=row, column=image_col_index)
                
                if image_cell.value and 'DISPIMG' in str(image_cell.value):
                    # 提取图片ID
                    formula = str(image_cell.value)
                    match = re.search(r'DISPIMG\("ID_([A-F0-9]+)",1\)', formula)
                    
                    if match:
                        image_id = match.group(1)
                        dispimg_mapping[row] = image_id
                        print(f"行{row}: 提取到图片ID: {image_id}")
            
            print(f"解析到 {len(dispimg_mapping)} 个DISPIMG公式")
            return dispimg_mapping
            
        except Exception as e:
            print(f"解析DISPIMG公式失败: {e}")
            return {}
    
    def _parse_dispimg_formula(self, formula):
        """解析DISPIMG公式，提取图片ID"""
        # 匹配DISPIMG公式中的图片ID，处理=_xlfn.DISPIMG("ID_图片ID",1)格式
        pattern = r'DISPIMG\("ID_([A-F0-9]+)",1\)'
        match = re.search(pattern, formula)
        if match:
            return match.group(1)
        return None
    
    def extract_all_images(self, output_dir: str = None) -> Dict[str, Any]:
        """
        提取Excel文件中的所有图片（仅使用WPS格式处理）
        
        Args:
            output_dir: 输出目录，如果为None则使用默认目录
            
        Returns:
            Dict[str, Any]: 提取结果，包含提取的图片信息和统计
        """
        if output_dir is None:
            output_dir = os.path.join(os.path.dirname(self.excel_file_path), 'extracted_images')
        
        os.makedirs(output_dir, exist_ok=True)
        
        try:
            # 复制Excel文件为zip格式
            zip_path = self.excel_file_path + '.zip'
            shutil.copy2(self.excel_file_path, zip_path)
            
            # 创建临时目录用于解压
            with tempfile.TemporaryDirectory() as temp_dir:
                extract_dir = os.path.join(temp_dir, 'extracted')
                
                # 解压Excel文件
                with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                    zip_ref.extractall(extract_dir)
                
                # 解析图片映射关系（仅使用WPS格式处理）
                print("开始解析WPS格式图片映射...")
                wps_mapping = self.parse_wps_cellimages_mapping(extract_dir)
                print(f"WPS格式映射解析完成，找到 {len(wps_mapping)} 个映射关系")
                if wps_mapping:
                    print(f"WPS映射示例: {dict(list(wps_mapping.items())[:2])}")
                
                # 提取图片文件
                extracted_images = self.extract_images_from_media(extract_dir, output_dir)
                
                # 解析DISPIMG公式
                dispimg_formulas = self.parse_dispimg_formulas(self.excel_file_path)
                print(f"解析到 {len(dispimg_formulas)} 个DISPIMG公式")
                
                # 建立正确的图片映射（仅使用WPS映射）
                correct_mapping = self.build_correct_image_mapping(extracted_images, dispimg_formulas, extract_dir)
                
                # 清理临时zip文件
                if os.path.exists(zip_path):
                    os.remove(zip_path)
                
                return {
                    'extracted_images': extracted_images,
                    'dispimg_formulas': dispimg_formulas,
                    'correct_mapping': correct_mapping,
                    'total_images': len(extracted_images),
                    'total_formulas': len(dispimg_formulas),
                    'output_dir': output_dir
                }
                
        except Exception as e:
            # 清理临时zip文件
            if os.path.exists(zip_path):
                os.remove(zip_path)
            raise Exception(f"提取图片失败: {e}")
    
    def build_correct_image_mapping(self, extracted_images: Dict[str, str], dispimg_mapping: Dict[int, str], extract_dir: str) -> Dict[str, str]:
        """
        建立正确的图片ID映射关系，为每个ID创建对应的图片副本
        
        Args:
            extracted_images: 提取的图片映射
            dispimg_mapping: DISPIMG公式映射
            extract_dir: 解压目录
            
        Returns:
            Dict[str, str]: 图片ID到文件路径的映射（ID已添加ID_前缀）
        """
        print(f"\n=== build_correct_image_mapping 开始 ===")
        print(f"extracted_images数量: {len(extracted_images)}")
        print(f"dispimg_mapping数量: {len(dispimg_mapping)}")
        
        correct_mapping = {}
        
        try:
            # 获取输出目录
            output_dir = os.path.dirname(list(extracted_images.values())[0]) if extracted_images else os.getcwd()
            
            # 建立图片文件到所有使用它的ID的映射
            file_to_ids = {}
            
            print(f"\n=== 分析映射关系 ===")
            print(f"提取的图片数量: {len(extracted_images)}")
            print(f"DISPIMG公式数量: {len(dispimg_mapping)}")
            
            # 首先获取WPS格式映射
            wps_mappings = {}
            try:
                wps_mappings = self.parse_wps_cellimages_mapping(extract_dir)
                print(f"WPS格式映射数量: {len(wps_mappings)}")
                print(f"WPS映射内容预览: {dict(list(wps_mappings.items())[:3])}")  # 显示前3个映射
            except Exception as e:
                print(f"WPS格式映射解析失败: {e}")
            
            # 建立DISPIMG公式中的图片ID与提取图片的对应关系
            for row, dispimg_image_id in dispimg_mapping.items():
                source_file = None
                
                # 使用WPS格式映射（主要方法）
                # 注意：WPS映射的键有ID_前缀，而DISPIMG公式中的ID没有前缀
                wps_image_id = f"ID_{dispimg_image_id}"
                if wps_image_id in wps_mappings:
                    wps_info = wps_mappings[wps_image_id]
                    actual_file = wps_info.get('actual_file', '')
                    if actual_file and actual_file in extracted_images:
                        source_file = extracted_images[actual_file]
                        print(f"行{row}: {wps_image_id} -> {os.path.basename(source_file)} (WPS映射)")
                
                if not source_file:
                    print(f"行{row}: ID_{dispimg_image_id} -> 未找到对应图片")
            
            # 建立文件到ID的反向映射（用于创建副本）
            for row, dispimg_image_id in dispimg_mapping.items():
                wps_image_id = f"ID_{dispimg_image_id}"
                if wps_image_id in wps_mappings:
                    wps_info = wps_mappings[wps_image_id]
                    actual_file = wps_info.get('actual_file', '')
                    # 注意：extracted_images的键是去掉扩展名的文件名
                    file_key = os.path.splitext(actual_file)[0] if actual_file else ''
                    if actual_file and file_key in extracted_images:
                        source_file = extracted_images[file_key]
                        full_image_id = wps_image_id  # 已经包含ID_前缀
                        
                        if source_file not in file_to_ids:
                            file_to_ids[source_file] = []
                        file_to_ids[source_file].append(full_image_id)
                        print(f"调试: 添加映射 {source_file} -> {full_image_id}")
                    else:
                        print(f"调试: ID {dispimg_image_id} 的实际文件 {actual_file} (键: {file_key}) 不在提取图片中")
                else:
                    print(f"调试: ID {dispimg_image_id} 不在WPS映射中")
            
            # 为每个ID创建对应的图片副本
            print(f"\n=== 创建ID图片副本 ===")
            print(f"输出目录: {output_dir}")
            
            for source_file, id_list in file_to_ids.items():
                if len(id_list) > 1:
                    print(f"\n图片文件 '{os.path.basename(source_file)}' 被 {len(id_list)} 个ID使用:")
                    for image_id in id_list:
                        print(f"  - {image_id}")
                else:
                    print(f"\n图片文件 '{os.path.basename(source_file)}' 被 1 个ID使用: {id_list[0]}")
                
                # 为每个ID创建副本
                for image_id in id_list:
                    # 创建目标文件名（添加ID_前缀）
                    file_ext = os.path.splitext(source_file)[1]
                    target_filename = f"{image_id}{file_ext}"
                    target_path = os.path.join(output_dir, target_filename)
                    
                    try:
                        # 复制图片文件
                        shutil.copy2(source_file, target_path)
                        correct_mapping[image_id] = target_path
                        print(f"  ✓ 创建副本: {image_id} -> {target_filename}")
                        
                    except Exception as e:
                        print(f"  ✗ 复制失败: {image_id} -> {e}")
            
            print(f"\n=== 副本创建完成 ===")
            print(f"成功创建 {len(correct_mapping)} 个ID图片副本")
            print(f"每个ID现在都有对应的图片文件，ID已添加ID_前缀")
            
            return correct_mapping
            
        except Exception as e:
            print(f"建立图片ID映射失败: {e}")
            # 返回原始的图片映射作为备选，并添加ID_前缀
            return {f"ID_{k}": v for k, v in extracted_images.items()}
    

    
    def parse_wps_cellimages_mapping(self, extract_dir: str) -> Dict[str, Dict[str, Any]]:
        """
        解析WPS格式的cellimages.xml映射关系（仅保留WPS格式特殊处理）
        
        Args:
            extract_dir: 解压目录
            
        Returns:
            Dict[str, Dict[str, Any]]: WPS格式的图片ID到映射信息的字典
        """
        wps_mappings = {}
        
        try:
            # 解析cellimages.xml
            cellimages_path = os.path.join(extract_dir, 'xl', 'cellimages.xml')
            
            print(f"检查cellimages.xml路径: {cellimages_path}")
            if not os.path.exists(cellimages_path):
                print("cellimages.xml文件不存在")
                return {}
            
            print("解析WPS格式cellimages.xml文件...")
            tree = ET.parse(cellimages_path)
            root = tree.getroot()
            
            # WPS特有命名空间
            namespaces = {
                'wps': 'http://www.wps.cn/officeDocument/2017/etCustomData',
                'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
            }
            
            # 查找WPS格式的cellImage元素
            cellimages_mapping = {}
            
            # 查找wps:cellImage元素
            for cellimage in root.findall('.//wps:cellImage', namespaces):
                # 查找xdr:pic元素
                pic = cellimage.find('.//xdr:pic', namespaces)
                if pic is not None:
                    # 查找cNvPr元素获取图片ID（存储在name属性中）
                    cNvPr = pic.find('.//xdr:cNvPr', namespaces)
                    if cNvPr is not None:
                        image_id = cNvPr.get('name')
                        if image_id:
                            # 查找blip元素获取embed_id
                            blip = pic.find('.//a:blip', namespaces)
                            if blip is not None:
                                embed_id = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                                if embed_id:
                                    cellimages_mapping[image_id] = embed_id
                                    print(f"✓ WPS格式映射: {image_id} -> {embed_id}")
            
            # 解析cellimages.xml.rels文件获取embed_id到实际文件的映射
            rels_path = os.path.join(extract_dir, 'xl', '_rels', 'cellimages.xml.rels')
            
            if os.path.exists(rels_path):
                print("解析WPS格式cellimages.xml.rels文件...")
                tree = ET.parse(rels_path)
                root = tree.getroot()
                
                # 解析关系映射
                embed_to_file = {}
                for rel in root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                    rel_id = rel.get('Id')
                    target = rel.get('Target')
                    if rel_id and target:
                        embed_to_file[rel_id] = target
                        print(f'WPS关系映射: {rel_id} -> {target}')
                
                # 建立WPS格式的完整映射关系
                for image_id, embed_id in cellimages_mapping.items():
                    if embed_id in embed_to_file:
                        file_path = embed_to_file[embed_id]
                        # 提取文件名
                        file_name = os.path.basename(file_path)
                        
                        # 获取media目录中的实际文件名
                        media_path = os.path.join(extract_dir, 'xl', 'media', file_name)
                        
                        if os.path.exists(media_path):
                            wps_mappings[image_id] = {
                                'image_id': image_id,
                                'embed_id': embed_id,
                                'actual_file': file_name,
                                'file_path': media_path,
                                'wps_format': True
                            }
                            
                            print(f'WPS格式完整映射: {image_id} -> {file_name}')
                        else:
                            print(f'WPS格式文件不存在: {file_name}')
                    else:
                        print(f'WPS格式未找到关系映射: {embed_id}')
                
                return wps_mappings
            else:
                print("WPS格式cellimages.xml.rels文件不存在")
                return {}
            
        except Exception as e:
            print(f"解析WPS格式cellimages映射失败: {e}")
            return {}

    def cleanup(self):
        """清理临时文件"""
        try:
            if os.path.exists(self.temp_dir):
                shutil.rmtree(self.temp_dir)
                print(f"临时文件已清理: {self.temp_dir}")
        except Exception as e:
            print(f"清理临时文件失败: {e}")


def main():
    """测试函数"""
    # 测试Excel图片提取
    excel_file = "智能家居模具库.xlsx"
    output_dir = "extracted_images"
    
    if os.path.exists(excel_file):
        extractor = ExcelImageExtractor(excel_file)
        try:
            images = extractor.extract_all_images(output_dir)
            print(f"测试完成: 提取了 {len(images)} 张图片")
        finally:
            extractor.cleanup()
    else:
        print(f"Excel文件不存在: {excel_file}")


if __name__ == "__main__":
    main()