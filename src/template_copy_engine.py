#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
模板复制引擎模块
负责复制模板格式，保持模板格式只复制文字和图片内容
"""

import os
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from typing import Dict, List, Any, Optional, Tuple
import shutil


class TemplateCopyEngine:
    """模板复制引擎类"""
    
    def __init__(self):
        """初始化模板复制引擎"""
        self.source_workbook: Optional[Workbook] = None
        self.source_worksheet: Optional[Worksheet] = None
        self.target_workbook: Optional[Workbook] = None
        self.target_worksheet: Optional[Worksheet] = None
        
    def copy_template(self, source_template_path: str, target_template_path: str, 
                     copy_data: bool = False) -> bool:
        """
        复制模板格式，保持模板格式只复制文字和图片内容
        
        Args:
            source_template_path: 源模板文件路径
            target_template_path: 目标模板文件路径
            copy_data: 是否复制数据内容（默认为False，只复制格式）
            
        Returns:
            bool: 是否复制成功
        """
        try:
            # 检查源文件是否存在
            if not os.path.exists(source_template_path):
                print(f"❌ 源模板文件不存在: {source_template_path}")
                return False
            
            # 检查文件格式
            if not source_template_path.lower().endswith('.xlsx'):
                print(f"❌ 源模板文件格式不正确，必须是.xlsx格式: {source_template_path}")
                return False
            
            print(f"🚀 开始复制模板...")
            print(f"   • 源模板: {source_template_path}")
            print(f"   • 目标模板: {target_template_path}")
            print(f"   • 复制数据: {'是' if copy_data else '否'}")
            
            # 加载源模板
            print("📋 加载源模板...")
            self.source_workbook = openpyxl.load_workbook(source_template_path)
            self.source_worksheet = self.source_workbook.active
            
            # 创建目标工作簿
            print("📄 创建目标模板...")
            self.target_workbook = Workbook()
            self.target_worksheet = self.target_workbook.active
            
            # 复制工作表名称
            self.target_worksheet.title = self.source_worksheet.title
            
            # 复制列宽
            self._copy_column_widths()
            
            # 复制行高
            self._copy_row_heights()
            
            # 复制单元格样式和内容
            self._copy_cell_styles_and_content(copy_data)
            
            # 复制合并单元格
            self._copy_merged_cells()
            
            # 复制图片
            self._copy_images()
            
            # 保存目标模板
            print("💾 保存目标模板...")
            self.target_workbook.save(target_template_path)
            
            # 关闭工作簿
            self.source_workbook.close()
            self.target_workbook.close()
            
            print("✅ 模板复制完成")
            return True
            
        except Exception as e:
            print(f"❌ 模板复制失败: {e}")
            if self.source_workbook:
                self.source_workbook.close()
            if self.target_workbook:
                self.target_workbook.close()
            return False
    
    def _copy_column_widths(self):
        """复制列宽"""
        if not self.source_worksheet or not self.target_worksheet:
            return
        
        print("   📏 复制列宽...")
        
        # 获取源工作表的所有列维度
        source_columns = self.source_worksheet.column_dimensions
        
        # 复制所有列的宽度设置，包括默认宽度
        for col_letter, source_dim in source_columns.items():
            # 复制列宽
            if source_dim.width:
                self.target_worksheet.column_dimensions[col_letter].width = source_dim.width
            else:
                # 如果没有设置宽度，使用默认宽度
                self.target_worksheet.column_dimensions[col_letter].width = 8.43  # Excel默认列宽
        
        # 确保复制所有可能存在的列（最大到Z列，即26列）
        max_cols = max(self.source_worksheet.max_column, 26)  # 至少复制26列
        for col in range(1, max_cols + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            if col_letter not in source_columns:
                # 对于没有显式设置宽度的列，确保有默认宽度
                if col_letter not in self.target_worksheet.column_dimensions:
                    self.target_worksheet.column_dimensions[col_letter].width = 8.43
        
        # 不再进行任何写死的列宽调整，完全按照模板的实际列宽复制
        # 这样用户手动调整模板列宽时，生成的采购清单也会相应调整
    
    def _copy_row_heights(self):
        """复制行高"""
        if not self.source_worksheet or not self.target_worksheet:
            return
        
        print("   📏 复制行高...")
        
        # 复制行高
        for row in range(1, self.source_worksheet.max_row + 1):
            row_dim = self.source_worksheet.row_dimensions.get(row)
            
            if row_dim and row_dim.height:
                self.target_worksheet.row_dimensions[row].height = row_dim.height
    
    def _copy_cell_styles_and_content(self, copy_data: bool):
        """复制单元格样式和内容"""
        if not self.source_worksheet or not self.target_worksheet:
            return
        
        print("   🎨 复制单元格样式和内容...")
        
        # 复制单元格样式和内容
        for row in range(1, self.source_worksheet.max_row + 1):
            for col in range(1, self.source_worksheet.max_column + 1):
                source_cell = self.source_worksheet.cell(row=row, column=col)
                target_cell = self.target_worksheet.cell(row=row, column=col)
                
                # 复制内容（如果允许复制数据）
                if copy_data:
                    target_cell.value = source_cell.value
                else:
                    # 只复制标题行和格式行
                    if row <= 2:  # 复制前两行（标题行）
                        target_cell.value = source_cell.value
                
                # 复制样式
                self._copy_cell_style(source_cell, target_cell)
    
    def _copy_cell_style(self, source_cell, target_cell):
        """复制单元格样式"""
        try:
            # 复制字体
            if source_cell.font:
                target_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic,
                    color=source_cell.font.color
                )
            
            # 复制填充
            if source_cell.fill:
                target_cell.fill = PatternFill(
                    fill_type=source_cell.fill.fill_type,
                    start_color=source_cell.fill.start_color,
                    end_color=source_cell.fill.end_color
                )
            
            # 复制边框
            if source_cell.border:
                border = Border(
                    left=Side(border_style=source_cell.border.left.border_style, 
                             color=source_cell.border.left.color) if source_cell.border.left else None,
                    right=Side(border_style=source_cell.border.right.border_style, 
                              color=source_cell.border.right.color) if source_cell.border.right else None,
                    top=Side(border_style=source_cell.border.top.border_style, 
                            color=source_cell.border.top.color) if source_cell.border.top else None,
                    bottom=Side(border_style=source_cell.border.bottom.border_style, 
                               color=source_cell.border.bottom.color) if source_cell.border.bottom else None
                )
                target_cell.border = border
            
            # 复制对齐方式
            if source_cell.alignment:
                target_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    wrap_text=source_cell.alignment.wrap_text,
                    shrink_to_fit=source_cell.alignment.shrink_to_fit,
                    indent=source_cell.alignment.indent
                )
                
        except Exception as e:
            # 样式复制失败不影响整体流程
            pass
    
    def _copy_merged_cells(self):
        """复制合并单元格"""
        if not self.source_worksheet or not self.target_worksheet:
            return
        
        print("   🔗 复制合并单元格...")
        
        # 复制合并单元格
        for merged_range in self.source_worksheet.merged_cells.ranges:
            self.target_worksheet.merge_cells(str(merged_range))
    
    def _copy_images(self):
        """复制图片"""
        if not self.source_worksheet or not self.target_worksheet:
            return
        
        print("   🖼️  复制图片...")
        
        # 复制图片
        for image in self.source_worksheet._images:
            try:
                # 创建图片副本
                img = Image(image.ref)
                img.anchor = image.anchor
                self.target_worksheet.add_image(img)
            except Exception as e:
                print(f"      ⚠️  复制图片失败: {e}")
    
    def create_enhanced_template(self, source_template_path: str, target_template_path: str,
                                pdid_data: Dict[str, Any]) -> bool:
        """
        创建增强采购清单模板
        
        Args:
            source_template_path: 源模板文件路径
            target_template_path: 目标模板文件路径
            pdid_data: PDID数据，包含产品信息和数量
            
        Returns:
            bool: 是否创建成功
        """
        try:
            print(f"🚀 开始创建增强采购清单...")
            print(f"   • 源模板: {source_template_path}")
            print(f"   • 目标清单: {target_template_path}")
            print(f"   • 产品数量: {len(pdid_data.get('products', []))}")
            
            # 复制模板格式
            if not self.copy_template(source_template_path, target_template_path, copy_data=False):
                return False
            
            # 重新加载目标模板进行数据填充
            print("📊 填充采购数据...")
            
            workbook = openpyxl.load_workbook(target_template_path)
            worksheet = workbook.active
            
            # 填充数据
            data_start_row = 2  # 从第2行开始填充数据
            
            for i, product in enumerate(pdid_data.get('products', [])):
                row = data_start_row + i
                
                # 填充产品信息
                worksheet.cell(row=row, column=1).value = product.get('设备品类', '')
                worksheet.cell(row=row, column=2).value = product.get('设备名称', '')
                worksheet.cell(row=row, column=3).value = product.get('品牌', '')
                worksheet.cell(row=row, column=4).value = product.get('型号', '')
                worksheet.cell(row=row, column=5).value = product.get('数量', 0)
                worksheet.cell(row=row, column=6).value = product.get('单位', '')
                worksheet.cell(row=row, column=7).value = product.get('单价', 0)
                worksheet.cell(row=row, column=8).value = product.get('小计', 0)
                worksheet.cell(row=row, column=9).value = product.get('产品图片', '')
                worksheet.cell(row=row, column=10).value = product.get('备注', '')
                worksheet.cell(row=row, column=11).value = product.get('产品链接', '')
            
            # 保存增强模板
            workbook.save(target_template_path)
            workbook.close()
            
            print("✅ 增强采购清单创建完成")
            return True
            
        except Exception as e:
            print(f"❌ 创建增强采购清单失败: {e}")
            return False


def test_template_copy():
    """测试模板复制功能"""
    print("=" * 60)
    print("🧪 模板复制引擎测试")
    print("=" * 60)
    
    # 测试文件路径
    source_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), '采购清单模板.xlsx')
    target_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), '测试复制模板.xlsx')
    
    if not os.path.exists(source_path):
        print(f"❌ 源模板文件不存在: {source_path}")
        return
    
    # 创建复制引擎
    engine = TemplateCopyEngine()
    
    # 测试模板复制
    success = engine.copy_template(source_path, target_path, copy_data=False)
    
    if success:
        print(f"✅ 模板复制测试成功")
        print(f"   • 源文件: {source_path}")
        print(f"   • 目标文件: {target_path}")
        
        # 检查文件是否创建成功
        if os.path.exists(target_path):
            file_size = os.path.getsize(target_path)
            print(f"   • 文件大小: {file_size} 字节")
            print("✅ 目标文件创建成功")
        else:
            print("❌ 目标文件创建失败")
    else:
        print("❌ 模板复制测试失败")


def test_enhanced_template():
    """测试增强模板创建功能"""
    print("=" * 60)
    print("🧪 增强模板创建测试")
    print("=" * 60)
    
    # 测试数据
    pdid_data = {
        'products': [
            {
                '设备品类': '智能开关',
                '设备名称': '二键智能开关',
                '品牌': '领普',
                '型号': 'KP2',
                '数量': 3,
                '单位': '个',
                '单价': 89.00,
                '小计': 267.00,
                '产品图片': '',
                '备注': 'PDID: 2',
                '产品链接': 'https://example.com/product/2'
            },
            {
                '设备品类': '智能开关',
                '设备名称': '四键智能开关',
                '品牌': '易来',
                '型号': 'K4',
                '数量': 4,
                '单位': '个',
                '单价': 109.00,
                '小计': 436.00,
                '产品图片': '',
                '备注': 'PDID: 8',
                '产品链接': 'https://example.com/product/8'
            }
        ]
    }
    
    # 测试文件路径
    source_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), '采购清单模板.xlsx')
    target_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), '增强采购清单.xlsx')
    
    if not os.path.exists(source_path):
        print(f"❌ 源模板文件不存在: {source_path}")
        return
    
    # 创建复制引擎
    engine = TemplateCopyEngine()
    
    # 测试增强模板创建
    success = engine.create_enhanced_template(source_path, target_path, pdid_data)
    
    if success:
        print(f"✅ 增强模板创建测试成功")
        print(f"   • 源文件: {source_path}")
        print(f"   • 目标文件: {target_path}")
        
        # 检查文件是否创建成功
        if os.path.exists(target_path):
            file_size = os.path.getsize(target_path)
            print(f"   • 文件大小: {file_size} 字节")
            print(f"   • 产品数量: {len(pdid_data['products'])}")
            print("✅ 增强采购清单创建成功")
        else:
            print("❌ 目标文件创建失败")
    else:
        print("❌ 增强模板创建测试失败")


if __name__ == "__main__":
    # 运行测试
    test_template_copy()
    print()
    test_enhanced_template()