#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
模板加载器模块
负责加载和验证采购清单模板文件，检查模板的列名和布局结构
"""

import os
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, List, Any, Optional, Tuple


class TemplateLoader:
    """模板加载器类"""
    
    def __init__(self):
        """初始化模板加载器"""
        self.workbook: Optional[Workbook] = None
        self.worksheet: Optional[Worksheet] = None
        self.template_info: Dict[str, Any] = {}
        
    def load_template(self, template_path: str) -> bool:
        """
        加载采购清单模板文件
        
        Args:
            template_path: 模板文件路径
            
        Returns:
            bool: 是否加载成功
        """
        try:
            # 检查文件是否存在
            if not os.path.exists(template_path):
                print(f"❌ 模板文件不存在: {template_path}")
                return False
            
            # 检查文件格式
            if not template_path.lower().endswith('.xlsx'):
                print(f"❌ 模板文件格式不正确，必须是.xlsx格式: {template_path}")
                return False
            
            # 加载模板文件
            print(f"🔍 加载模板文件: {template_path}")
            self.workbook = openpyxl.load_workbook(template_path)
            
            # 获取活动工作表
            self.worksheet = self.workbook.active
            
            # 分析模板结构
            if not self._analyze_template_structure():
                print("❌ 模板结构分析失败")
                return False
            
            print("✅ 模板文件加载成功")
            return True
            
        except Exception as e:
            print(f"❌ 加载模板文件失败: {e}")
            return False
    
    def _analyze_template_structure(self) -> bool:
        """
        分析模板结构
        
        Returns:
            bool: 是否分析成功
        """
        try:
            if not self.worksheet:
                return False
            
            # 获取模板基本信息
            self.template_info = {
                'sheet_name': self.worksheet.title,
                'max_row': self.worksheet.max_row,
                'max_column': self.worksheet.max_column,
                'column_names': [],
                'merged_cells': [],
                'data_start_row': 1,
                'data_end_row': self.worksheet.max_row
            }
            
            # 获取列名（假设第一行是标题行）
            if self.worksheet.max_row >= 1:
                for col in range(1, self.worksheet.max_column + 1):
                    cell_value = self.worksheet.cell(row=1, column=col).value
                    if cell_value:
                        self.template_info['column_names'].append({
                            'column': col,
                            'name': str(cell_value),
                            'letter': openpyxl.utils.get_column_letter(col)
                        })
            
            # 获取合并单元格信息
            for merged_range in self.worksheet.merged_cells.ranges:
                self.template_info['merged_cells'].append({
                    'range': str(merged_range),
                    'min_row': merged_range.min_row,
                    'max_row': merged_range.max_row,
                    'min_col': merged_range.min_col,
                    'max_col': merged_range.max_col
                })
            
            # 查找数据区域
            self._find_data_region()
            
            print(f"📊 模板结构分析完成:")
            print(f"   • 工作表: {self.template_info['sheet_name']}")
            print(f"   • 行数: {self.template_info['max_row']}")
            print(f"   • 列数: {self.template_info['max_column']}")
            print(f"   • 列名: {[col['name'] for col in self.template_info['column_names']]}")
            print(f"   • 合并单元格: {len(self.template_info['merged_cells'])}个")
            
            return True
            
        except Exception as e:
            print(f"❌ 分析模板结构失败: {e}")
            return False
    
    def _find_data_region(self):
        """查找数据区域"""
        if not self.worksheet:
            return
        
        # 查找数据开始行（跳过标题行）
        data_start_row = 2  # 默认从第2行开始
        
        # 查找数据结束行（最后一个有数据的行）
        data_end_row = self.worksheet.max_row
        
        # 反向查找，找到最后一个有数据的行
        for row in range(self.worksheet.max_row, 0, -1):
            has_data = False
            for col in range(1, self.worksheet.max_column + 1):
                cell_value = self.worksheet.cell(row=row, column=col).value
                if cell_value:
                    has_data = True
                    break
            if has_data:
                data_end_row = row
                break
        
        self.template_info['data_start_row'] = data_start_row
        self.template_info['data_end_row'] = data_end_row
        
        print(f"   • 数据区域: 第{data_start_row}行到第{data_end_row}行")
    
    def validate_template(self) -> Tuple[bool, List[str]]:
        """
        验证模板格式
        
        Returns:
            Tuple[bool, List[str]]: (是否验证通过, 错误信息列表)
        """
        errors = []
        
        if not self.worksheet:
            errors.append("模板未加载")
            return False, errors
        
        # 检查必要的列名
        required_columns = ['设备品类', '设备名称', '品牌', '型号', '数量', '单位', '单价', '小计']
        existing_columns = [col['name'] for col in self.template_info['column_names']]
        
        for required_col in required_columns:
            if required_col not in existing_columns:
                errors.append(f"缺少必要列: {required_col}")
        
        # 检查数据区域
        if self.template_info['data_start_row'] >= self.template_info['data_end_row']:
            errors.append("数据区域无效")
        
        # 检查是否有足够的行用于数据填充
        available_rows = self.template_info['data_end_row'] - self.template_info['data_start_row'] + 1
        if available_rows < 10:  # 至少需要10行用于数据填充
            errors.append(f"数据区域行数不足，当前只有{available_rows}行")
        
        if errors:
            return False, errors
        else:
            return True, []
    
    def get_template_info(self) -> Dict[str, Any]:
        """
        获取模板信息
        
        Returns:
            Dict[str, Any]: 模板信息
        """
        return self.template_info.copy()
    
    def get_worksheet(self) -> Optional[Worksheet]:
        """
        获取工作表对象
        
        Returns:
            Optional[Worksheet]: 工作表对象
        """
        return self.worksheet
    
    def get_workbook(self) -> Optional[Workbook]:
        """
        获取工作簿对象
        
        Returns:
            Optional[Workbook]: 工作簿对象
        """
        return self.workbook
    
    def close(self):
        """关闭模板文件"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.worksheet = None


def load_and_validate_template(template_path: str) -> Tuple[bool, Optional[TemplateLoader], List[str]]:
    """
    加载并验证模板文件的便捷函数
    
    Args:
        template_path: 模板文件路径
        
    Returns:
        Tuple[bool, Optional[TemplateLoader], List[str]]: (是否成功, 模板加载器对象, 错误信息)
    """
    loader = TemplateLoader()
    
    # 加载模板
    if not loader.load_template(template_path):
        return False, None, ["模板加载失败"]
    
    # 验证模板
    is_valid, errors = loader.validate_template()
    
    if not is_valid:
        loader.close()
        return False, None, errors
    
    return True, loader, []


def main():
    """主函数 - 测试模板加载器"""
    print("=" * 60)
    print("📋 模板加载器测试")
    print("=" * 60)
    
    # 测试默认模板
    template_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), '采购清单模板.xlsx')
    
    if not os.path.exists(template_path):
        print(f"❌ 模板文件不存在: {template_path}")
        return
    
    # 加载并验证模板
    success, loader, errors = load_and_validate_template(template_path)
    
    if success:
        print("\n✅ 模板验证通过")
        
        # 显示模板信息
        template_info = loader.get_template_info()
        print(f"\n📊 模板详细信息:")
        print(f"   • 工作表名称: {template_info['sheet_name']}")
        print(f"   • 总行数: {template_info['max_row']}")
        print(f"   • 总列数: {template_info['max_column']}")
        print(f"   • 数据区域: 第{template_info['data_start_row']}行到第{template_info['data_end_row']}行")
        print(f"   • 列名列表:")
        for col in template_info['column_names']:
            print(f"     - {col['letter']}列: {col['name']}")
        print(f"   • 合并单元格数量: {len(template_info['merged_cells'])}")
        
        loader.close()
    else:
        print(f"\n❌ 模板验证失败:")
        for error in errors:
            print(f"   • {error}")


if __name__ == "__main__":
    main()