#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
检查模板文件格式和样式
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def check_template_format():
    """检查模板文件的格式和样式"""
    
    print("🔍 检查采购清单模板格式和样式...")
    
    try:
        # 加载模板文件
        workbook = openpyxl.load_workbook('../采购清单模板.xlsx')
        worksheet = workbook.active
        
        print(f"✅ 模板文件加载成功，工作表: {worksheet.title}")
        print(f"   数据范围: A1:{worksheet.max_column}{worksheet.max_row}")
        
        # 检查标题行样式
        print("\n🎨 标题行样式检查:")
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            if cell.value:
                print(f"   列{col}({cell.value}):")
                if cell.fill:
                    print(f"     填充颜色: {cell.fill.fgColor.rgb if cell.fill.fgColor else '无'}")
                if cell.font:
                    print(f"     字体: {cell.font.name}, 大小: {cell.font.size}, 颜色: {cell.font.color.rgb if cell.font.color else '默认'}")
                if cell.alignment:
                    print(f"     对齐: 水平{cell.alignment.horizontal}, 垂直{cell.alignment.vertical}, 换行: {cell.alignment.wrapText}")
        
        # 检查行高和列宽
        print("\n📏 行高和列宽检查:")
        print(f"   标题行高度: {worksheet.row_dimensions[1].height}")
        for col in range(1, min(12, worksheet.max_column + 1)):
            col_letter = openpyxl.utils.get_column_letter(col)
            width = worksheet.column_dimensions[col_letter].width
            print(f"   列{col_letter}宽度: {width}")
        
        # 检查是否有图片
        print("\n🖼️  图片检查:")
        image_count = len(worksheet._images)
        print(f"   图片数量: {image_count}")
        
        # 检查单元格边框
        print("\n📐 边框样式检查:")
        sample_cell = worksheet['A1']
        if sample_cell.border:
            border = sample_cell.border
            print(f"   边框样式: 左{border.left.style}, 右{border.right.style}, 上{border.top.style}, 下{border.bottom.style}")
        
        print("\n🎉 模板格式检查完成！")
        
        return workbook, worksheet
        
    except Exception as e:
        print(f"❌ 检查过程中出错: {e}")
        return None, None

def main():
    """主函数"""
    print("=" * 50)
    print("📋 采购清单模板格式分析")
    print("=" * 50)
    
    workbook, worksheet = check_template_format()
    
    if workbook:
        print("\n✅ 模板格式分析完成，可以基于此模板生成采购清单")
        workbook.close()

if __name__ == "__main__":
    main()