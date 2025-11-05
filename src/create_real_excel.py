#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
创建真实的智能家居模具库Excel文件
"""

import openpyxl
from openpyxl.styles import Font, Alignment

def create_real_excel_file():
    """创建真实的Excel文件"""
    
    # 创建Excel工作簿
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "智能家居模具库"
    
    # 设置表头
    headers = [
        "设备品类", "设备名称", "设备简称", "是否启用", 
        "单价", "品牌", "主规格", "单位", 
        "渠道", "采购链接", "设备图片"
    ]
    
    # 设置表头样式
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 智能开关数据
    switch_data = [
        # 格式: [设备品类, 设备名称, 设备简称, 是否启用, 单价, 品牌, 主规格, 单位, 渠道, 采购链接, 设备图片]
        ["智能开关", "一键智能开关", "一键开关", "是", 79, "颜工", "86型", "个", "电商", "https://example.com/switch1", "assets/images/switches/一键.png"],
        ["智能开关", "二键智能开关", "二键开关", "是", 89, "颜工", "86型", "个", "电商", "https://example.com/switch2", "assets/images/switches/二键.png"],
        ["智能开关", "三键智能开关", "三键开关", "是", 99, "颜工", "86型", "个", "电商", "https://example.com/switch3", "assets/images/switches/三键.png"],
        ["智能开关", "四键智能开关", "四键开关", "是", 119, "颜工", "86型", "个", "电商", "https://example.com/switch4", "assets/images/switches/四键.png"]
    ]
    
    # 写入数据
    for row, data in enumerate(switch_data, 2):
        for col, value in enumerate(data, 1):
            sheet.cell(row=row, column=col, value=value)
    
    # 设置列宽
    column_widths = {
        'A': 12,  # 设备品类
        'B': 15,  # 设备名称
        'C': 10,  # 设备简称
        'D': 8,   # 是否启用
        'E': 8,   # 单价
        'F': 10,  # 品牌
        'G': 12,  # 主规格
        'H': 6,   # 单位
        'I': 8,   # 渠道
        'J': 20,  # 采购链接
        'K': 30   # 设备图片
    }
    
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width
    
    # 保存文件
    excel_path = "../智能家居模具库.xlsx"
    workbook.save(excel_path)
    print(f"✅ 真实的Excel文件已创建: {excel_path}")
    print(f"📊 包含 {len(switch_data)} 个智能开关产品")
    print("📁 设备图片使用您存放的本地图片路径")
    
    return excel_path

def generate_ppt_from_real_excel():
    """从真实的Excel文件生成PPT模具库"""
    
    from excel_to_ppt_converter import ExcelToPPTConverter
    
    # 创建转换器实例
    converter = ExcelToPPTConverter(image_folder="../assets/images")
    
    # Excel文件路径
    excel_path = "../智能家居模具库.xlsx"
    ppt_path = "../智能家居模具库.pptx"
    
    # 生成PPT
    print("\n🎨 开始生成PPT模具库...")
    success = converter.generate_ppt_from_excel(excel_path, ppt_path)
    
    if success:
        print(f"✅ PPT模具库生成成功: {ppt_path}")
        print("📋 包含智能开关一键到四键的完整模具")
        print("🖼️ 使用您存放的本地开关图片")
    else:
        print("❌ PPT生成失败")
    
    return success

if __name__ == "__main__":
    # 创建真实的Excel文件
    excel_path = create_real_excel_file()
    
    # 从Excel生成PPT
    generate_ppt_from_real_excel()
    
    print("\n🎉 所有任务完成！")
    print("📁 文件位置:")
    print(f"   - Excel文件: {excel_path}")
    print(f"   - PPT文件: ../智能家居模具库.pptx")
    print(f"   - 图片目录: ../assets/images/switches/")
    print("\n💡 您现在可以直接使用这个真实的模具库了！")