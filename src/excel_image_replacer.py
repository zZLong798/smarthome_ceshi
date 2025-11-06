"""
Excel图片替换器模块 - 将Excel中的WPS特定函数=DISPIMG(...)替换为真正嵌入单元格的本地图片
"""

import json
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage


class ExcelImageReplacer:
    """Excel图片替换器类"""
    
    def __init__(self, image_mapping_path=None):
        """
        初始化图片替换器
        
        Args:
            image_mapping_path: 图片映射JSON文件路径（如果为None，则使用默认路径）
        """
        if image_mapping_path is None:
            # 设置默认路径：项目根目录下的images/image_mapping.json
            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            self.image_mapping_path = os.path.join(base_dir, "images", "image_mapping.json")
        else:
            self.image_mapping_path = image_mapping_path
            
        self.image_mapping = self._load_image_mapping()
        
        # 图片插入配置
        self.target_img_height = 100  # 目标图片高度（像素）
        self.target_row_height = 80   # 目标行高（磅，1像素 ≈ 0.75磅）
        self.target_col_width = 25    # 目标列宽（字符）
    
    def _load_image_mapping(self):
        """加载图片映射文件"""
        try:
            with open(self.image_mapping_path, 'r', encoding='utf-8') as f:
                mapping_data = json.load(f)
            
            # 从mapping_relationships中提取PDID到图片路径的映射
            mapping_relationships = mapping_data.get('mapping_relationships', [])
            pdid_to_image_map = {}
            
            for mapping in mapping_relationships:
                product_id = mapping.get('product_id', '')
                real_image_file = mapping.get('real_image_file', '')
                
                if product_id and real_image_file:
                    # 处理相对路径，转换为绝对路径
                    if not os.path.isabs(real_image_file):
                        # 假设图片文件相对于项目根目录
                        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                        real_image_file = os.path.join(base_dir, real_image_file)
                    
                    pdid_to_image_map[product_id] = real_image_file
            
            print(f"✅ 成功加载图片映射，共 {len(pdid_to_image_map)} 个映射关系")
            return pdid_to_image_map
            
        except FileNotFoundError:
            print(f"❌ 映射文件未找到: {self.image_mapping_path}")
            return {}
        except json.JSONDecodeError:
            print(f"❌ 映射文件格式错误: {self.image_mapping_path}")
            return {}
        except Exception as e:
            print(f"❌ 加载映射文件失败: {e}")
            return {}
    
    def replace_dispimg_formulas(self, excel_path, output_path=None, 
                                pdid_column="A", image_column="I", 
                                start_row=2):
        """
        替换Excel中的DISPIMG公式为嵌入图片
        
        Args:
            excel_path: 输入Excel文件路径
            output_path: 输出Excel文件路径（如果为None，则覆盖原文件）
            pdid_column: PDID所在列
            image_column: 图片所在列
            start_row: 数据开始行（第1行通常是标题）
            
        Returns:
            bool: 是否替换成功
        """
        if not self.image_mapping:
            print("❌ 映射为空，停止处理")
            return False
        
        if output_path is None:
            output_path = excel_path.replace('.xlsx', '_with_images.xlsx')
        
        try:
            print(f"📂 正在打开Excel文件: {excel_path}")
            wb = load_workbook(excel_path)
            ws = wb.active
            
            print("🔄 开始处理行数据并嵌入图片...")
            
            # 设置图片列的宽度
            ws.column_dimensions[image_column].width = self.target_col_width
            
            # 遍历所有数据行
            processed_count = 0
            success_count = 0
            
            for row_num in range(start_row, ws.max_row + 1):
                # 获取PDID单元格和图片单元格
                pdid_cell = ws[f"{pdid_column}{row_num}"]
                image_cell = ws[f"{image_column}{row_num}"]
                
                pdid = str(pdid_cell.value).strip() if pdid_cell.value else None
                
                if not pdid:
                    print(f"   ⚠️ 第 {row_num} 行：PDID为空，跳过")
                    continue
                
                processed_count += 1
                
                # 查找图片文件路径
                image_path = self.image_mapping.get(pdid)
                
                if not image_path:
                    print(f"   ❌ 第 {row_num} 行：未找到PDID '{pdid}' 的映射图片")
                    image_cell.value = "未找到映射"  # 清空旧公式
                    continue
                
                # 检查图片文件是否存在
                if not os.path.exists(image_path):
                    print(f"   ❌ 第 {row_num} 行：图片文件不存在: {image_path}")
                    image_cell.value = "图片文件丢失"  # 清空旧公式
                    continue
                
                # 核心操作：替换DISPIMG公式为嵌入图片
                try:
                    # 1. 清空旧单元格内容（即=DISPIMG(...)公式）
                    image_cell.value = None
                    
                    # 2. 设置行高以容纳图片
                    ws.row_dimensions[row_num].height = self.target_row_height
                    
                    # 3. 加载图片
                    img = Image(image_path)
                    
                    # 4. 调整图片大小以适应单元格
                    # 保持宽高比，固定高度
                    original_img = PILImage.open(image_path)
                    scale = self.target_img_height / original_img.height
                    img.height = self.target_img_height
                    img.width = int(original_img.width * scale)
                    
                    # 5. 添加图片到工作表，锚定到单元格
                    ws.add_image(img, image_cell.coordinate)
                    
                    success_count += 1
                    print(f"   ✅ 第 {row_num} 行：成功嵌入图片 {os.path.basename(image_path)} 到 {image_cell.coordinate}")
                    
                except Exception as e:
                    print(f"   ❌ 第 {row_num} 行：插入图片 {image_path} 时出错: {e}")
                    image_cell.value = "图片插入失败"
            
            # 保存修改后的Excel
            print(f"\n💾 所有图片处理完毕，正在保存到: {output_path}")
            try:
                wb.save(output_path)
                wb.close()
                
                print(f"✅ 保存成功！")
                print(f"📊 处理统计：")
                print(f"   - 总处理行数: {processed_count}")
                print(f"   - 成功嵌入图片: {success_count}")
                print(f"   - 失败行数: {processed_count - success_count}")
                
                return True
                
            except Exception as e:
                print(f"❌ 保存文件失败: {e}")
                return False
                
        except FileNotFoundError:
            print(f"❌ Excel文件未找到: {excel_path}")
            return False
        except Exception as e:
            print(f"❌ 处理Excel文件失败: {e}")
            return False
    
    def batch_replace_excel_files(self, excel_directory, output_directory=None, 
                                 file_pattern="*.xlsx", **kwargs):
        """
        批量替换目录中所有Excel文件的DISPIMG公式
        
        Args:
            excel_directory: Excel文件目录
            output_directory: 输出目录（如果为None，则在原目录创建_with_images文件）
            file_pattern: 文件匹配模式
            **kwargs: 传递给replace_dispimg_formulas的其他参数
            
        Returns:
            Dict[str, bool]: 每个文件的处理结果
        """
        import glob
        
        if output_directory is None:
            output_directory = excel_directory
        
        # 确保输出目录存在
        os.makedirs(output_directory, exist_ok=True)
        
        # 查找所有Excel文件
        excel_files = glob.glob(os.path.join(excel_directory, file_pattern))
        
        if not excel_files:
            print(f"❌ 在目录 {excel_directory} 中未找到匹配 {file_pattern} 的Excel文件")
            return {}
        
        results = {}
        
        print(f"📁 开始批量处理目录 {excel_directory} 中的 {len(excel_files)} 个Excel文件")
        
        for excel_file in excel_files:
            filename = os.path.basename(excel_file)
            
            if output_directory == excel_directory:
                # 在原目录创建_with_images文件
                output_file = excel_file.replace('.xlsx', '_with_images.xlsx')
            else:
                # 在输出目录创建同名文件
                output_file = os.path.join(output_directory, filename)
            
            print(f"\n📄 处理文件: {filename}")
            
            success = self.replace_dispimg_formulas(excel_file, output_file, **kwargs)
            results[excel_file] = success
        
        # 统计结果
        success_count = sum(1 for result in results.values() if result)
        total_count = len(results)
        
        print(f"\n📊 批量处理完成:")
        print(f"   - 总处理文件数: {total_count}")
        print(f"   - 成功处理文件: {success_count}")
        print(f"   - 失败文件数: {total_count - success_count}")
        
        return results


def test_excel_image_replacer():
    """测试Excel图片替换器"""
    print("🧪 测试Excel图片替换器...")
    
    # 创建替换器实例
    replacer = ExcelImageReplacer()
    
    # 测试单个文件替换
    test_excel_path = "../智能设备采购清单.xlsx"  # 假设有这个文件
    
    if os.path.exists(test_excel_path):
        success = replacer.replace_dispimg_formulas(
            excel_path=test_excel_path,
            pdid_column="A",  # PDID在A列
            image_column="I",  # 图片在I列
            start_row=2       # 从第2行开始（第1行是标题）
        )
        
        if success:
            print("✅ Excel图片替换测试成功")
        else:
            print("❌ Excel图片替换测试失败")
    else:
        print(f"⚠️ 测试文件不存在: {test_excel_path}")
        print("💡 请先运行采购清单生成器生成测试文件")


if __name__ == "__main__":
    test_excel_image_replacer()