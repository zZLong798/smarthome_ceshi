"""
增强版Excel图片映射解析器
基于parse_correct_mapping_final.py，按照《Excel图片获取映射规则文档.md》实现完整映射链
"""

import os
import zipfile
import xml.etree.ElementTree as ET
import openpyxl
import re
from typing import Dict, List, Any, Optional


class EnhancedExcelImageMapper:
    """增强版Excel图片映射解析器"""
    
    def __init__(self, excel_path: str):
        """
        初始化映射解析器
        
        Args:
            excel_path: Excel文件路径
        """
        self.excel_path = excel_path
        self.temp_dir = "temp_excel_extract"
        
    def parse_enhanced_mapping(self) -> Dict[str, Dict[str, Any]]:
        """
        解析增强的图片映射关系
        
        Returns:
            包含完整映射链的字典
        """
        print("=== 开始解析增强版Excel图片映射 ===")
        
        # 1. 提取DISPIMG公式和PDID信息
        dispimg_mappings = self.extract_dispimg_formulas()
        print(f"✓ 提取到 {len(dispimg_mappings)} 个DISPIMG公式")
        
        # 2. 解析cellimages.xml获取完整映射关系
        cellimages_mappings = self.parse_cellimages_mapping()
        print(f"✓ 解析到 {len(cellimages_mappings)} 个cellimages映射")
        
        # 3. 建立完整映射链
        complete_mappings = self.build_complete_mapping_chain(dispimg_mappings, cellimages_mappings)
        print(f"✓ 建立 {len(complete_mappings)} 个完整映射链")
        
        # 4. 验证映射链完整性
        validation_result = self.validate_mapping_chain(complete_mappings)
        
        return complete_mappings
    
    def extract_dispimg_formulas(self) -> Dict[int, Dict[str, Any]]:
        """
        提取DISPIMG公式和PDID信息
        
        Returns:
            行号到DISPIMG信息的映射
        """
        try:
            workbook = openpyxl.load_workbook(self.excel_path)
            sheet = workbook.active
            
            dispimg_mappings = {}
            
            # 找到设备图片列
            image_col_index = None
            for col in range(1, sheet.max_column + 1):
                header = sheet.cell(row=1, column=col).value
                if header and '图片' in str(header):
                    image_col_index = col
                    print(f"找到设备图片列: 第{col}列 - {header}")
                    break
            
            if not image_col_index:
                print("未找到设备图片列")
                return {}
            
            # 分析所有行的DISPIMG公式
            for row in range(2, sheet.max_row + 1):
                # 获取PDID（第1列）
                pdid_cell = sheet.cell(row=row, column=1)
                pdid = pdid_cell.value if pdid_cell.value else ""
                
                # 获取设备简称（第4列）
                device_short_cell = sheet.cell(row=row, column=4)
                device_short = device_short_cell.value if device_short_cell.value else ""
                
                # 获取图片单元格
                image_cell = sheet.cell(row=row, column=image_col_index)
                
                if image_cell.value and 'DISPIMG' in str(image_cell.value):
                    # 提取图片ID
                    formula = str(image_cell.value)
                    # 增强正则表达式，支持更多WPS格式变体
                    # 标准格式: DISPIMG("ID_...",1)
                    # WPS格式1: =_xlfn.DISPIMG("ID_...",1)
                    # WPS格式2: =DISPIMG("ID_...",1)
                    # WPS格式3: DISPIMG("ID_...", 1) (带空格)
                    match = re.search(r'(?:=_?_?xlfn\.)?DISPIMG\s*\(\s*"([^"]+)"\s*,\s*\d+\s*\)', formula)
                    
                    if match:
                        image_id = match.group(1)
                        
                        # 生成正确的单元格引用（L列，第12列）
                        cell_reference = f"L{row}"
                        
                        dispimg_mappings[row] = {
                            'pdid': str(pdid).strip() if pdid else "",
                            'device_name': str(device_short).strip() if device_short else "",
                            'dispimg_formula': formula.strip(),
                            'image_id': image_id,
                            'row_number': row,
                            'cell_reference': cell_reference
                        }
                        
                        print(f"行{row}: PDID={pdid}, 设备简称={device_short}, DISPIMG图片ID={image_id}, 单元格={cell_reference}")
                    else:
                        print(f"行{row}: 无法解析DISPIMG公式: {formula}")
            
            workbook.close()
            return dispimg_mappings
            
        except Exception as e:
            print(f"提取DISPIMG公式失败: {e}")
            return {}
    
    def _extract_dispimg_formulas(self, sheet_data: Dict) -> List[Dict[str, str]]:
        """
        提取工作表中的DISPIMG公式
        
        Args:
            sheet_data: 工作表数据
            
        Returns:
            List[Dict[str, str]]: DISPIMG公式列表
        """
        dispimg_formulas = []
        
        try:
            # 获取单元格数据
            cells = sheet_data.get('cells', {})
            
            # 遍历所有单元格，查找DISPIMG公式
            for cell_ref, cell_data in cells.items():
                if not isinstance(cell_data, dict):
                    continue
                    
                # 获取单元格公式
                formula = cell_data.get('f', '')
                if not formula:
                    continue
                
                # 检查是否为DISPIMG公式（支持标准格式和WPS格式）
                # 标准格式: DISPIMG("图片ID",1)
                # WPS格式1: =_xlfn.DISPIMG("图片ID",1)
                # WPS格式2: =DISPIMG("图片ID",1)
                # WPS格式3: DISPIMG("图片ID", 1) (带空格)
                if 'DISPIMG' in formula:
                    # 增强正则表达式，支持更多WPS格式变体
                    match = re.search(r'(?:=_?_?xlfn\.)?DISPIMG\s*\(\s*"([^"]+)"\s*,\s*\d+\s*\)', formula)
                    if match:
                        image_id = match.group(1)
                        dispimg_formulas.append({
                            'cell_reference': cell_ref,
                            'image_id': image_id,
                            'formula': formula
                        })
                        print(f"发现DISPIMG公式: {cell_ref} -> {image_id}")
                    else:
                        print(f"无法解析DISPIMG公式: {formula}")
            
            return dispimg_formulas
            
        except Exception as e:
            print(f"提取DISPIMG公式失败: {e}")
            return []
    
    def parse_cellimages_mapping(self) -> Dict[str, Dict[str, Any]]:
        """
        解析cellimages.xml获取完整映射关系
        
        Returns:
            图片ID到完整映射信息的字典
        """
        try:
            # 解压Excel文件
            if not self._extract_excel():
                print("Excel文件解压失败")
                return {}
            
            # 解析cellimages.xml
            cellimages_path = os.path.join(self.temp_dir, 'xl', 'cellimages.xml')
            
            if not os.path.exists(cellimages_path):
                print("cellimages.xml文件不存在")
                return {}
            
            print("解析cellimages.xml文件...")
            tree = ET.parse(cellimages_path)
            root = tree.getroot()
            
            # 解析图片定义
            image_mappings = {}
            
            # 定义命名空间
            namespaces = {
                'wps': 'http://www.wps.cn/officeDocument/2017/etCustomData',
                'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
            }
            
            # 查找WPS格式的cellImage元素
            cellimages_mapping = {}
            
            # 查找wps:cellImage元素
            for cellimage in root.findall('.//wps:cellImage', namespaces):
                # 查找xdr:pic元素
                pic = cellimage.find('.//xdr:pic', namespaces)
                if pic is not None:
                    # 查找cNvPr元素获取图片ID（存储在name属性中）
                    cNvPr = pic.find('.//xdr:cNvPr', namespaces)
                    if cNvPr is not None:
                        image_id = cNvPr.get('name')
                        if image_id:
                            # 查找blip元素获取embed_id
                            blip = pic.find('.//a:blip', namespaces)
                            if blip is not None:
                                embed_id = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                                if embed_id:
                                    cellimages_mapping[image_id] = embed_id
                                    print(f"✓ 找到图片映射: {image_id} -> {embed_id}")
            
            # 如果没有找到WPS格式的映射，尝试标准格式
            if not cellimages_mapping:
                print("未找到WPS格式映射，尝试标准格式...")
                # 标准格式解析逻辑
                for cellimage in root.findall('.//cellImage'):
                    image_id = cellimage.get('id')
                    embed_id = cellimage.get('embed')
                    if image_id and embed_id:
                        cellimages_mapping[image_id] = embed_id
                        print(f"✓ 找到图片映射: {image_id} -> {embed_id}")
            
            # 解析cellimages.xml.rels文件获取embed_id到实际文件的映射
            rels_path = os.path.join(self.temp_dir, 'xl', '_rels', 'cellimages.xml.rels')
            
            if os.path.exists(rels_path):
                print("解析cellimages.xml.rels文件获取实际文件映射...")
                tree = ET.parse(rels_path)
                root = tree.getroot()
                
                # 解析关系映射
                embed_to_file = {}
                for rel in root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                    rel_id = rel.get('Id')
                    target = rel.get('Target')
                    if rel_id and target:
                        embed_to_file[rel_id] = target
                        print(f'关系ID: {rel_id} -> 文件: {target}')
                
                # 建立完整的映射关系
                complete_mappings = {}
                
                for image_id, embed_id in cellimages_mapping.items():
                    if embed_id in embed_to_file:
                        file_path = embed_to_file[embed_id]
                        # 提取文件名
                        file_name = os.path.basename(file_path)
                        
                        # 获取media目录中的实际文件名
                        media_path = os.path.join(self.temp_dir, 'xl', 'media', file_name)
                        
                        if os.path.exists(media_path):
                            mapping = {
                                'image_id': image_id,
                                'embed_id': embed_id,
                                'actual_file': file_name,
                                'file_path': media_path,
                                'r_id': embed_id,
                                'name': '',
                                'description': '',
                                'index': len(complete_mappings) + 1
                            }
                            complete_mappings[image_id] = mapping
                            
                            print(f'图片ID: {image_id}, 实际文件: {file_name}')
                        else:
                            print(f'图片ID: {image_id}, 文件不存在: {file_name}')
                    else:
                        print(f'图片ID: {image_id}, 未找到对应的关系映射')
                
                return complete_mappings
            else:
                print("cellimages.xml.rels文件不存在")
                return {}
            
        except Exception as e:
            print(f"解析cellimages映射失败: {e}")
            return {}
    
    def build_complete_mapping_chain(self, 
                                   dispimg_mappings: Dict[int, Dict], 
                                   cellimages_mappings: Dict[str, Dict]) -> Dict[str, Dict[str, Any]]:
        """
        建立完整映射链：PDID → DISPIMG公式 → 图片ID → rId → 图片文件
        
        Args:
            dispimg_mappings: DISPIMG公式映射
            cellimages_mappings: cellimages映射
            
        Returns:
            完整映射链字典
        """
        complete_mappings = {}
        
        for row_num, dispimg_info in dispimg_mappings.items():
            image_id = dispimg_info['image_id']
            pdid = dispimg_info['pdid']
            
            # 尝试匹配cellimages映射
            # DISPIMG公式中的image_id不带ID_前缀，而cellimages_mappings的键带ID_前缀
            wps_image_id = f"ID_{image_id}"
            cellimages_info = None
            
            # 首先尝试直接匹配
            if image_id in cellimages_mappings:
                cellimages_info = cellimages_mappings[image_id]
            # 然后尝试添加ID_前缀匹配
            elif wps_image_id in cellimages_mappings:
                cellimages_info = cellimages_mappings[wps_image_id]
                # 使用带ID_前缀的图片ID
                image_id = wps_image_id
            
            if cellimages_info:
                # 建立完整映射链
                mapping_key = f"pdid_{pdid}" if pdid else f"row_{row_num}"
                
                complete_mappings[mapping_key] = {
                    'pdid': pdid,
                    'device_name': dispimg_info['device_name'],
                    'row_number': row_num,
                    'dispimg_formula': dispimg_info['dispimg_formula'],
                    'image_id': image_id,  # 使用匹配后的image_id（可能已添加ID_前缀）
                    'r_id': cellimages_info.get('r_id', ''),
                    'actual_file': cellimages_info.get('actual_file', ''),
                    'file_path': cellimages_info.get('file_path', ''),
                    'description': cellimages_info.get('description', ''),
                    'cell_reference': f"L{row_num}",  # 使用L列（第12列）作为图片单元格引用
                    'validation_status': 'complete' if pdid and cellimages_info.get('actual_file') else 'incomplete'
                }
                
                print(f"建立映射链: PDID={pdid} -> {image_id} -> {cellimages_info.get('actual_file', '')}")
            else:
                print(f"✗ 未找到图片ID {image_id} (或 {wps_image_id}) 对应的cellimages映射")
        
        return complete_mappings
    
    def validate_mapping_chain(self, mapping_data: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
        """
        验证映射链完整性
        
        Args:
            mapping_data: 映射数据
            
        Returns:
            验证结果
        """
        print("\n=== 验证映射链完整性 ===")
        
        total_mappings = len(mapping_data)
        complete_mappings = 0
        missing_pdid = 0
        missing_images = 0
        
        for mapping_key, mapping_info in mapping_data.items():
            # 检查PDID是否存在
            if not mapping_info['pdid']:
                missing_pdid += 1
                mapping_info['validation_status'] = 'missing_pdid'
                print(f"缺失 {mapping_key}: 缺失PDID")
            # 检查图片文件是否存在
            elif not mapping_info.get('actual_file'):
                missing_images += 1
                mapping_info['validation_status'] = 'missing_image'
                print(f"缺失 {mapping_key}: 缺失图片文件")
            else:
                complete_mappings += 1
                mapping_info['validation_status'] = 'complete'
                print(f"完整 {mapping_key}: 映射完整")
        
        validation_summary = {
            'total_mappings': total_mappings,
            'complete_mappings': complete_mappings,
            'missing_pdid': missing_pdid,
            'missing_images': missing_images,
            'completeness_rate': round(complete_mappings / total_mappings * 100, 2) if total_mappings > 0 else 0
        }
        
        print(f"\n验证结果:")
        print(f"总映射数: {total_mappings}")
        print(f"完整映射: {complete_mappings}")
        print(f"缺失PDID: {missing_pdid}")
        print(f"缺失图片: {missing_images}")
        print(f"完整率: {validation_summary['completeness_rate']}%")
        
        return validation_summary
    
    def _extract_excel(self) -> bool:
        """解压Excel文件到临时目录"""
        try:
            # 清理临时目录
            if os.path.exists(self.temp_dir):
                import shutil
                shutil.rmtree(self.temp_dir)
            
            # 创建临时目录
            os.makedirs(self.temp_dir, exist_ok=True)
            
            # 解压Excel文件
            with zipfile.ZipFile(self.excel_path, 'r') as zip_ref:
                zip_ref.extractall(self.temp_dir)
            
            print(f"Excel文件解压成功: {self.temp_dir}")
            return True
            
        except Exception as e:
            print(f"Excel文件解压失败: {e}")
            return False
    
    def cleanup(self):
        """清理临时文件"""
        try:
            if os.path.exists(self.temp_dir):
                import shutil
                shutil.rmtree(self.temp_dir)
                print("临时文件清理完成")
        except Exception as e:
            print(f"清理临时文件失败: {e}")


def parse_enhanced_excel_image_mapping(excel_path: str) -> Dict[str, Any]:
    """
    解析增强的Excel图片映射关系（对外接口）
    
    Args:
        excel_path: Excel文件路径
        
    Returns:
        包含完整映射链和验证结果的字典
    """
    mapper = EnhancedExcelImageMapper(excel_path)
    
    try:
        # 解析映射关系
        mapping_data = mapper.parse_enhanced_mapping()
        
        # 验证映射链
        validation_result = mapper.validate_mapping_chain(mapping_data)
        
        result = {
            'mapping_chain': mapping_data,
            'validation_summary': validation_result,
            'success': True,
            'message': '映射解析完成'
        }
        
        return result
        
    except Exception as e:
        return {
            'mapping_chain': {},
            'validation_summary': {},
            'success': False,
            'message': f'映射解析失败: {e}'
        }
    
    finally:
        mapper.cleanup()


if __name__ == "__main__":
    # 测试增强映射解析器
    excel_file = "智能家居模具库.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"Excel文件不存在: {excel_file}")
    else:
        print("=== 测试增强版Excel图片映射解析器 ===")
        
        result = parse_enhanced_excel_image_mapping(excel_file)
        
        if result['success']:
            print("\n=== 映射解析成功 ===")
            print(f"总映射数: {result['validation_summary']['total_mappings']}")
            print(f"完整率: {result['validation_summary']['completeness_rate']}%")
            
            # 保存结果到文件
            import json
            with open('enhanced_image_mapping.json', 'w', encoding='utf-8') as f:
                json.dump(result, f, ensure_ascii=False, indent=2)
            print("映射结果已保存到 enhanced_image_mapping.json")
        else:
            print(f"映射解析失败: {result['message']}")